# Improve Airpot Code
from openpyxl import Workbook
import openpyxl
import time

class Contacts():
    def __init__(self):
        self.totalContacts = []
        self.route = 'C:\\Users\David\Desktop\Aeropuerto\\'
        self.fileRoute = self.route + 'myContacts.xlsx'
    
    def addContact(self):
        newContact = {
            "name" :"--",
            "lastName": "--",
            "age": 0,
            "phone": "--"
        }
        name        = input("Enter Contact Name: ")
        lastName    = input("Enter Contact Last Name: ")
        age         = input("Enter Contact Age: ")
        phone       = input("Enter Contact Phone: ")
        if len(lastName) > 0:
            newContact["lastName"] = lastName
        if len(name) > 0:
            newContact["name"] = name
        if len(age) > 0:
            newContact["age"] = age
        if len(phone) > 0:
            newContact["phone"] = phone
        self.totalContacts.append(newContact)

    def printContacts(self):
        a = "{}.- {} {} \tAge: {} Phone: {}"
        j = 1
        for x in self.totalContacts:
            print(a.format(j, x["lastName"], x["name"], x["age"], x["phone"]))
            j += 1

    def updateContact(self):
        updatedContact = {
            "name" :"--",
            "lastName": "--",
            "age": 0,
            "phone": "--"
        }
        self.printContacts()
        contactNumber = int(input("Please Enter Contact Number: "))
        name        = input("Enter Contact Name: ")
        lastName    = input("Enter Contact Last Name: ")
        age         = input("Enter Contact Age: ")
        phone       = input("Enter Contact Phone: ")
        if len(lastName) > 0:
            updatedContact["lastName"] = lastName
        if len(name) > 0:
            updatedContact["name"] = name
        if len(age) > 0:
            updatedContact["age"] = age
        if len(phone) > 0:
            updatedContact["phone"] = phone
        self.totalContacts[contactNumber-1] = updatedContact
        print("\n\nContact Updated Successfully!\n")

    def deleteContact(self):
        self.printContacts()
        number = int(input("Please Enter Contact Number: "))
        self.totalContacts.remove(self.totalContacts[number-1])
        print("\n\nContact Deleted Successfully! \n")
        self.printContacts()

    def deleteEverything(self):
        self.totalContacts.clear()
        print("\n\nContact List Deleted Sucessfully!\n")

    def createExcel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Contacts"
        ws['B1'] = 'Last Name'
        ws['C1'] = 'Name'
        ws['D1'] = 'Age'
        ws['E1'] = 'Phone'
        wb.save(self.fileRoute)

    def loadExcel(self):
        wb = openpyxl.load_workbook(self.fileRoute)
        i = 2
        ws = wb["Contacts"]
        while(ws['A' + str(i)].value != None):
            excelLine = {
                "name" :"--",
                "lastName": "--",
                "age": 0,
                "phone": "--"
            }
            if len(str(ws['B' + str(i)].value)) > 0:
                excelLine["lastName"] = ws['B' + str(i)].value
            if len(str(ws['C' + str(i)].value)) > 0:
                excelLine["name"] = ws['C' + str(i)].value
            if len(str(ws['D' + str(i)].value)) > 0:
                excelLine["age"] = ws['D' + str(i)].value
            if len(str(ws['E' + str(i)].value)) > 0:
                excelLine["phone"] = ws['E' + str(i)].value
            self.totalContacts.append(excelLine)
            i += 1
            print("Contact Added Successfully!")
        print('\n\n')
        user.printContacts()

    def updateExcel(self):
        wb = openpyxl.load_workbook(self.fileRoute)
        i = 2
        ws = wb["Contacts"]
        for x in self.totalContacts:
            ws['A' + str(i)] = str(i-1)
            ws['B' + str(i)] = x["lastName"]
            ws['C' + str(i)] = x["name"]
            ws['D' + str(i)] = x["age"]
            ws['E' + str(i)] = x["phone"]
            i+=1
        wb.save(self.fileRoute)
        print('\n\nExcel Successfully Updated!\n')

    def goBackToMenu(self):
        input("\n\nPress Any Key To Go Back To the Menu....")


def startMenu():
    menu = "1.-Add Contact\n2.-Edit Contact\n3.-Delete Contact\n4.-View All\n5.-Load Excel\n6.-Update Excel\n7.-Create Excel(TBD)\n8.-Exit"
    print("\n"*5)
    print("="*25)
    print(menu)
    print("="*25)
    option= input("Enter Option: ")
    return option

user = Contacts()
action = int(startMenu())
while action != 8:
    if action == 1:
        user.addContact()
        user.goBackToMenu()
    if action == 2:
        user.updateContact()
        user.goBackToMenu()
    if action == 3:
        user.deleteContact()
        user.goBackToMenu()
    if action == 4:
        user.printContacts()
        user.goBackToMenu()
    if action == 5:
        user.loadExcel()
        user.goBackToMenu()
    if action == 6:
        user.updateExcel()
        user.goBackToMenu()
    if action == 7:
        user.createExcel()
        user.goBackToMenu()
    action = int(startMenu())

print("GoodBye")